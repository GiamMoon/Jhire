from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session
from fastapi.responses import FileResponse
import pandas as pd
import os
from ...infrastructure.database import get_db
from ...infrastructure.models import Order, Product, OrderItem

router = APIRouter()

@router.get("/excel")
def export_excel(db: Session = Depends(get_db)):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter
    
    orders = db.query(Order).order_by(Order.created_at.desc()).all()
    
    data = []
    for order in orders:
        data.append({
            "Código": f"JHIRE-{order.id}",
            "Estado": order.status.upper(),
            "Monto (S/)": round(float(order.total_price), 2),
            "Fecha": order.created_at.strftime("%d/%m/%Y"),
            "Hora": order.created_at.strftime("%H:%M")
        })
    
    df = pd.DataFrame(data)
    filepath = "informe_corporativo_jhire.xlsx"
    
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # Sheet 1: Data
        df.to_excel(writer, sheet_name='Detalle Ventas', index=False, startrow=3)
        ws = writer.sheets['Detalle Ventas']
        
        # Header branding
        ws.merge_cells('A1:E1')
        ws['A1'] = 'JHIRE S.A.C. — Informe Ejecutivo de Ventas 2026'
        ws['A1'].font = Font(name='Calibri', bold=True, size=16, color='003461')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells('A2:E2')
        ws['A2'] = f'Total de transacciones analizadas: {len(orders)} | Generado automáticamente por JHIRE ERP'
        ws['A2'].font = Font(name='Calibri', size=10, color='727781', italic=True)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Style headers (row 4 since data starts at row 4)
        header_fill = PatternFill(start_color='003461', end_color='003461', fill_type='solid')
        header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        thin_border = Border(
            left=Side(style='thin', color='C2C7D1'),
            right=Side(style='thin', color='C2C7D1'),
            top=Side(style='thin', color='C2C7D1'),
            bottom=Side(style='thin', color='C2C7D1')
        )
        
        for col_idx in range(1, 6):
            cell = ws.cell(row=4, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # Style data rows (zebra stripes)
        light_fill = PatternFill(start_color='F2F4FF', end_color='F2F4FF', fill_type='solid')
        for row_idx in range(5, 5 + len(orders)):
            for col_idx in range(1, 6):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.font = Font(name='Calibri', size=10)
                cell.alignment = Alignment(horizontal='center')
                if row_idx % 2 == 0:
                    cell.fill = light_fill
        
        # Total row
        total_row = 5 + len(orders)
        ws.cell(row=total_row, column=2, value='TOTAL INGRESOS:').font = Font(name='Calibri', bold=True, size=12, color='003461')
        ws.cell(row=total_row, column=3, value=round(sum(float(o.total_price) for o in orders), 2)).font = Font(name='Calibri', bold=True, size=12, color='003461')
        ws.cell(row=total_row, column=3).number_format = '#,##0.00'
        
        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 10
        
        # --- PIE CHART: Orders by Status ---
        status_counts = {}
        for o in orders:
            status_counts[o.status] = status_counts.get(o.status, 0) + 1
        
        # Write status data in hidden area for chart
        chart_start_col = 7
        ws.cell(row=4, column=chart_start_col, value='Estado').font = Font(bold=True, size=9, color='727781')
        ws.cell(row=4, column=chart_start_col + 1, value='Cantidad').font = Font(bold=True, size=9, color='727781')
        
        for i, (status, count) in enumerate(status_counts.items()):
            ws.cell(row=5 + i, column=chart_start_col, value=status)
            ws.cell(row=5 + i, column=chart_start_col + 1, value=count)
        
        pie = PieChart()
        pie.title = "Distribución por Estado"
        pie.style = 10
        labels = Reference(ws, min_col=chart_start_col, min_row=5, max_row=4 + len(status_counts))
        vals = Reference(ws, min_col=chart_start_col + 1, min_row=4, max_row=4 + len(status_counts))
        pie.add_data(vals, titles_from_data=True)
        pie.set_categories(labels)
        pie.width = 18
        pie.height = 12
        ws.add_chart(pie, "G1")
        
        # --- BAR CHART: Top 10 Orders by Value ---
        bar = BarChart()
        bar.type = "col"
        bar.title = "Top 10 Órdenes por Valor (S/)"
        bar.style = 10
        bar.y_axis.title = "Monto S/"
        bar.x_axis.title = "Orden"
        
        top_orders = sorted(orders, key=lambda o: float(o.total_price), reverse=True)[:10]
        bar_start_col = 10
        ws.cell(row=4, column=bar_start_col, value='Orden').font = Font(bold=True, size=9, color='727781')
        ws.cell(row=4, column=bar_start_col + 1, value='Monto').font = Font(bold=True, size=9, color='727781')
        
        for i, o in enumerate(top_orders):
            ws.cell(row=5 + i, column=bar_start_col, value=f"#{o.id}")
            ws.cell(row=5 + i, column=bar_start_col + 1, value=round(float(o.total_price), 2))
        
        bar_labels = Reference(ws, min_col=bar_start_col, min_row=5, max_row=4 + len(top_orders))
        bar_vals = Reference(ws, min_col=bar_start_col + 1, min_row=4, max_row=4 + len(top_orders))
        bar.add_data(bar_vals, titles_from_data=True)
        bar.set_categories(bar_labels)
        bar.shape = 4
        bar.width = 20
        bar.height = 12
        ws.add_chart(bar, "G16")
    
    return FileResponse(filepath, filename="informe_corporativo_jhire.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@router.get("/pdf")
def export_pdf(db: Session = Depends(get_db)):
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.graphics.shapes import Drawing, Rect, String
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    
    orders = db.query(Order).order_by(Order.created_at.desc()).limit(200).all()
    filepath = "informe_ejecutivo_jhire.pdf"
    
    doc = SimpleDocTemplate(filepath, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=40, bottomMargin=50)
    story = []
    styles = getSampleStyleSheet()
    
    # Custom Styles
    title_style = ParagraphStyle(
        name="TitleCustom",
        parent=styles['Heading1'],
        fontName="Helvetica-Bold",
        fontSize=22,
        alignment=1,
        spaceAfter=5,
        textColor=colors.HexColor("#003461")
    )
    
    subtitle_style = ParagraphStyle(
        name="SubTitleCustom",
        parent=styles['Normal'],
        fontName="Helvetica",
        fontSize=10,
        alignment=1,
        spaceAfter=25,
        textColor=colors.HexColor("#727781")
    )
    
    section_style = ParagraphStyle(
        name="SectionTitle",
        parent=styles['Heading3'],
        fontName="Helvetica-Bold",
        fontSize=13,
        spaceBefore=20,
        spaceAfter=10,
        textColor=colors.HexColor("#003461"),
        borderPadding=(0, 0, 5, 0),
    )
    
    # --- HEADER ---
    # Decorative line
    header_drawing = Drawing(535, 4)
    header_drawing.add(Rect(0, 0, 535, 4, fillColor=colors.HexColor("#003461"), strokeColor=None))
    story.append(header_drawing)
    story.append(Spacer(1, 10))
    
    story.append(Paragraph("INFORME EJECUTIVO DE VENTAS", title_style))
    story.append(Paragraph("JHIRE S.A.C. — Sistema ERP Inteligente 2026", subtitle_style))
    
    # --- KPI SUMMARY BOXES ---
    total_rev = sum(float(o.total_price) for o in orders)
    completed = sum(1 for o in orders if o.status == "Completado")
    pending = sum(1 for o in orders if o.status == "En Proceso")
    cancelled = sum(1 for o in orders if o.status == "Cancelado")
    
    kpi_data = [
        ["MÉTRICA", "VALOR"],
        ["Total Órdenes Analizadas", str(len(orders))],
        ["Ingresos Totales", f"S/ {total_rev:,.2f}"],
        ["Ticket Promedio", f"S/ {(total_rev/max(1,len(orders))):,.2f}"],
        ["Completadas / Pendientes / Canceladas", f"{completed} / {pending} / {cancelled}"]
    ]
    
    kpi_table = Table(kpi_data, colWidths=[280, 250])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#003461")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f4ff")]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#c2c7d1")),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 15))
    
    # --- PIE CHART ---
    story.append(Paragraph("DISTRIBUCIÓN POR ESTADO DE ÓRDENES", section_style))
    
    status_counts = {}
    for o in orders:
        status_counts[o.status] = status_counts.get(o.status, 0) + 1
    
    if status_counts:
        drawing = Drawing(400, 180)
        pie = Pie()
        pie.x = 100
        pie.y = 10
        pie.width = 150
        pie.height = 150
        pie.data = list(status_counts.values())
        pie.labels = [f"{k} ({v})" for k, v in status_counts.items()]
        
        palette = [
            colors.HexColor("#003461"),
            colors.HexColor("#16a34a"),
            colors.HexColor("#ba1a1a"),
            colors.HexColor("#727781"),
            colors.HexColor("#6366f1")
        ]
        for i in range(len(pie.data)):
            pie.slices[i].fillColor = palette[i % len(palette)]
            pie.slices[i].strokeWidth = 0.5
            pie.slices[i].strokeColor = colors.white
        
        pie.sideLabels = 1
        pie.slices.fontName = 'Helvetica'
        pie.slices.fontSize = 9
        drawing.add(pie)
        story.append(drawing)
    
    story.append(Spacer(1, 10))
    
    # --- BAR CHART: Top 10 ---
    top_orders = sorted(orders, key=lambda o: float(o.total_price), reverse=True)[:10]
    if top_orders:
        story.append(Paragraph("TOP 10 ÓRDENES POR VALOR", section_style))
        
        bar_drawing = Drawing(500, 180)
        bc = VerticalBarChart()
        bc.x = 50
        bc.y = 10
        bc.height = 140
        bc.width = 420
        bc.data = [[float(o.total_price) for o in top_orders]]
        bc.categoryAxis.categoryNames = [f"#{o.id}" for o in top_orders]
        bc.categoryAxis.labels.fontSize = 8
        bc.categoryAxis.labels.fontName = 'Helvetica'
        bc.valueAxis.valueMin = 0
        bc.valueAxis.labels.fontSize = 8
        bc.valueAxis.labels.fontName = 'Helvetica'
        bc.bars[0].fillColor = colors.HexColor("#003461")
        bc.bars[0].strokeColor = None
        bar_drawing.add(bc)
        story.append(bar_drawing)
    
    story.append(Spacer(1, 15))
    
    # --- DETAILED TABLE ---
    story.append(Paragraph("DETALLE DE TRANSACCIONES", section_style))
    
    table_data = [["#", "CÓDIGO", "ESTADO", "MONTO (S/)", "FECHA"]]
    for idx, o in enumerate(orders[:50], 1):
        table_data.append([
            str(idx),
            f"JHIRE-{o.id}",
            o.status.upper(),
            f"S/ {float(o.total_price):,.2f}",
            o.created_at.strftime('%d/%m/%Y')
        ])
    
    # Total
    table_data.append(["", "", "TOTAL:", f"S/ {total_rev:,.2f}", ""])
    
    t = Table(table_data, colWidths=[30, 80, 120, 140, 100])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#003461")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor("#f2f4ff")]),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.HexColor("#c2c7d1")),
        
        # Total row
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 11),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#eaedff")),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor("#003461")),
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor("#003461")),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(t)
    
    # Footer
    story.append(Spacer(1, 25))
    footer_drawing = Drawing(535, 4)
    footer_drawing.add(Rect(0, 0, 535, 2, fillColor=colors.HexColor("#c2c7d1"), strokeColor=None))
    story.append(footer_drawing)
    story.append(Spacer(1, 5))
    
    footer_style = ParagraphStyle(
        name="Footer",
        parent=styles['Normal'],
        fontName="Helvetica",
        fontSize=8,
        alignment=1,
        textColor=colors.HexColor("#727781")
    )
    story.append(Paragraph("Documento generado automáticamente por JHIRE ERP 2026 · Confidencial", footer_style))
    story.append(Paragraph("Av. Industrial 1234, Lima, Perú · contacto@jhire.com.pe · RUC: 20123456789", footer_style))
    
    doc.build(story)
    
    return FileResponse(filepath, filename="informe_ejecutivo_jhire.pdf", media_type="application/pdf")
